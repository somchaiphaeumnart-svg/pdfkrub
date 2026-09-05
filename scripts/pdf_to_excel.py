#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
High-quality Advanced PDF to Excel (.xlsx) converter.
Extracts tables and tabular data from PDF files with intelligent layout detection,
custom page ranges, worksheet structuring (single sheet vs multi-sheet),
and native numeric formatting.

Engine hierarchy:
1. pdfplumber (best table grid structure detection: Lattice & Stream) + openpyxl
2. PyMuPDF (fitz) page.find_tables() + openpyxl or built-in XLSX writer
3. pdftotext -layout (system poppler-utils) fallback + built-in XLSX writer
"""

import sys
import os
import re
import argparse
import zipfile
import xml.sax.saxutils as saxutils
import subprocess

def col_to_letter(col_idx):
    """Convert 0-indexed column number to Excel column name (0 -> A, 1 -> B, 26 -> AA)."""
    result = ""
    col_idx += 1
    while col_idx > 0:
        col_idx, remainder = divmod(col_idx - 1, 26)
        result = chr(65 + remainder) + result
    return result

import unicodedata
import tempfile
import csv

def clean_thai_text(text):
    """Normalize Thai Unicode and remove corrupt CID glyph placeholders."""
    if not text:
        return ""
    # Unicode normalize NFC
    text = unicodedata.normalize('NFC', str(text))
    # Remove CID glyph artifacts like (cid:123)
    text = re.sub(r'\(cid:\d+\)', '', text)
    # Fix Thai sara am: nikhahit (U+0E4D) + sara aa (U+0E32) -> sara am (U+0E33)
    text = text.replace('\u0e4d\u0e32', '\u0e33')
    # Strip non-printable ASCII control characters except newline and tab, and replacement char
    text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F\ufffd]', '', text)
    return text.strip()

def clean_cell_value(val):
    if val is None:
        return ""
    return clean_thai_text(val)

def parse_page_ranges(page_spec, total_pages):
    """
    Parse page specifications such as 'all', '1,3-5,8', etc.
    Returns 0-indexed page list.
    """
    if not page_spec or page_spec.strip().lower() == 'all':
        return list(range(total_pages))

    pages = set()
    parts = page_spec.replace(';', ',').split(',')
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if '-' in part:
            bounds = part.split('-')
            if len(bounds) == 2:
                try:
                    start = int(bounds[0].strip())
                    end = int(bounds[1].strip())
                    for p in range(max(1, start), min(total_pages, end) + 1):
                        pages.add(p - 1)
                except ValueError:
                    pass
        else:
            try:
                p = int(part)
                if 1 <= p <= total_pages:
                    pages.add(p - 1)
            except ValueError:
                pass

    sorted_pages = sorted(list(pages))
    return sorted_pages if sorted_pages else list(range(total_pages))

def write_xlsx_builtin(pages_data, output_path, sheet_mode="single"):
    """Generates standard OpenXML .xlsx file using only Python built-in modules."""
    if not pages_data:
        pages_data = [[[""]]]

    # Prepare sheet structure based on sheet_mode
    if sheet_mode == "single" and len(pages_data) > 1:
        consolidated = []
        for p_idx, page_rows in enumerate(pages_data):
            if p_idx > 0 and consolidated and page_rows:
                consolidated.append([]) # blank spacer row
            consolidated.extend(page_rows)
        final_sheets = [consolidated]
        sheet_titles = ["Sheet1"]
    else:
        final_sheets = pages_data
        sheet_titles = [f"Page {i+1}" if len(pages_data) > 1 else "Sheet1" for i in range(len(pages_data))]

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    temp_output = output_path + ".tmp"

    with zipfile.ZipFile(temp_output, 'w', zipfile.ZIP_DEFLATED) as zf:
        num_sheets = len(final_sheets)

        # [Content_Types].xml
        content_types = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
            '  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
            '  <Default Extension="xml" ContentType="application/xml"/>',
            '  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
            '  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>'
        ]
        for i in range(num_sheets):
            content_types.append(f'  <Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
        content_types.append('</Types>')
        zf.writestr('[Content_Types].xml', '\n'.join(content_types))

        # _rels/.rels
        zf.writestr('_rels/.rels',
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
            '  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>\n'
            '</Relationships>'
        )

        # xl/_rels/workbook.xml.rels
        wb_rels = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        ]
        for i in range(num_sheets):
            wb_rels.append(f'  <Relationship Id="rId{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{i+1}.xml"/>')
        wb_rels.append(f'  <Relationship Id="rId{num_sheets+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>')
        wb_rels.append('</Relationships>')
        zf.writestr('xl/_rels/workbook.xml.rels', '\n'.join(wb_rels))

        # xl/workbook.xml
        wb_xml = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
            '  <sheets>'
        ]
        for i in range(num_sheets):
            wb_xml.append(f'    <sheet name="{sheet_titles[i]}" sheetId="{i+1}" r:id="rId{i+1}"/>')
        wb_xml.append('  </sheets>')
        wb_xml.append('</workbook>')
        zf.writestr('xl/workbook.xml', '\n'.join(wb_xml))

        # xl/styles.xml
        zf.writestr('xl/styles.xml',
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">\n'
            '  <fonts count="2">\n'
            '    <font><sz val="11"/><name val="Calibri"/></font>\n'
            '    <font><b/><sz val="11"/><name val="Calibri"/></font>\n'
            '  </fonts>\n'
            '  <fills count="3">\n'
            '    <fill><patternFill patternType="none"/></fill>\n'
            '    <fill><patternFill patternType="gray125"/></fill>\n'
            '    <fill><patternFill patternType="solid"><fgColor rgb="FFF1F5F9"/></patternFill></fill>\n'
            '  </fills>\n'
            '  <borders count="2">\n'
            '    <border><left/><right/><top/><bottom/><diagonal/></border>\n'
            '    <border><left style="thin"><color rgb="FFCBD5E1"/></left><right style="thin"><color rgb="FFCBD5E1"/></right><top style="thin"><color rgb="FFCBD5E1"/></top><bottom style="thin"><color rgb="FFCBD5E1"/></bottom></border>\n'
            '  </borders>\n'
            '  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>\n'
            '  <cellXfs count="2">\n'
            '    <xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/>\n'
            '    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>\n'
            '  </cellXfs>\n'
            '</styleSheet>'
        )

        # worksheets/sheet{i+1}.xml
        for i, page_rows in enumerate(final_sheets):
            sheet_xml = [
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
                '  <sheetData>'
            ]
            row_counter = 1
            for r_idx, row in enumerate(page_rows):
                row_cells = []
                is_header = (r_idx == 0 and len(row) > 1)
                for col_idx, cell_val in enumerate(row):
                    str_val = clean_cell_value(cell_val)
                    if not str_val:
                        continue
                    cell_ref = f"{col_to_letter(col_idx)}{row_counter}"
                    # Check if numeric
                    is_num = False
                    try:
                        clean_num = str_val.replace(',', '')
                        float(clean_num)
                        if str_val == '0' or not str_val.startswith('0') or str_val.startswith('0.'):
                            is_num = True
                            str_val = clean_num
                    except ValueError:
                        is_num = False

                    style_attr = ' s="1"' if is_header else ''
                    if is_num:
                        row_cells.append(f'<c r="{cell_ref}"{style_attr}><v>{str_val}</v></c>')
                    else:
                        escaped = saxutils.escape(str_val)
                        row_cells.append(f'<c r="{cell_ref}"{style_attr} t="inlineStr"><is><t>{escaped}</t></is></c>')

                if row_cells:
                    sheet_xml.append(f'    <row r="{row_counter}">')
                    sheet_xml.extend(f'      {c}' for c in row_cells)
                    sheet_xml.append('    </row>')
                row_counter += 1

            sheet_xml.append('  </sheetData>')
            sheet_xml.append('</worksheet>')
            zf.writestr(f'xl/worksheets/sheet{i+1}.xml', '\n'.join(sheet_xml))

    os.replace(temp_output, output_path)
    return True

def save_to_excel(pages_data, output_path, sheet_mode="single"):
    """Save extracted pages rows to XLSX file using openpyxl with high-fidelity formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Prepare sheets list
        if sheet_mode == "single" and len(pages_data) > 1:
            consolidated = []
            for p_idx, page_rows in enumerate(pages_data):
                if p_idx > 0 and consolidated and page_rows:
                    consolidated.append([]) # blank separator
                consolidated.extend(page_rows)
            final_sheets = [consolidated]
            sheet_titles = ["Sheet1"]
        else:
            final_sheets = pages_data
            sheet_titles = [f"Page {i+1}" if len(pages_data) > 1 else "Sheet1" for i in range(len(pages_data))]

        header_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        header_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        data_font = Font(name="Calibri", size=11, color="0F172A")

        for s_idx, page_rows in enumerate(final_sheets):
            ws = wb.create_sheet(title=sheet_titles[s_idx])
            ws.views.sheetView[0].showGridLines = True

            col_widths = {}
            for r_idx, row in enumerate(page_rows):
                cleaned_row = []
                for cell in row:
                    val = clean_cell_value(cell)
                    # Convert to numeric if applicable
                    try:
                        clean_num = val.replace(',', '')
                        if val == '0' or not val.startswith('0') or val.startswith('0.'):
                            float_val = float(clean_num)
                            cleaned_row.append(float_val if '.' in clean_num else int(clean_num))
                            continue
                    except Exception:
                        pass
                    cleaned_row.append(val)

                ws.append(cleaned_row)
                current_row_idx = ws.max_row
                is_first_row = (r_idx == 0 and len(cleaned_row) > 1)

                for c_idx, cell_value in enumerate(cleaned_row, 1):
                    cell = ws.cell(row=current_row_idx, column=c_idx)
                    cell.border = thin_border
                    if is_first_row:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.font = data_font
                        if isinstance(cell_value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")

                    # Track column width
                    str_len = len(str(cell_value or ''))
                    col_widths[c_idx] = max(col_widths.get(c_idx, 10), min(str_len + 3, 50))

            # Apply column widths
            for col_idx, width in col_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = max(width, 12)

        wb.save(output_path)
        return True
    except ImportError:
        pass
    except Exception as e:
        sys.stderr.write(f"openpyxl warning, falling back to built-in writer: {e}\n")

    return write_xlsx_builtin(pages_data, output_path, sheet_mode=sheet_mode)

def is_text_corrupted(pages_data):
    """
    Check if extracted text suffers from font encoding corruption,
    missing ToUnicode maps ((cid:X)), or scrambled glyphs.
    """
    total_cells = 0
    corrupted_count = 0

    for page in pages_data:
        for row in page:
            for cell in row:
                s = str(cell).strip()
                if not s:
                    continue
                total_cells += 1
                # 1. Direct CID glyph placeholder check
                if '(cid:' in s or '\ufffd' in s:
                    corrupted_count += 3
                    continue
                # 2. Check for reversed date like 6202/ or 5202/ or 4652/ (reversed 2564)
                if re.search(r'\b(6202|5202|4202|7652|6652|5652|4652)/\d+', s):
                    corrupted_count += 3
                    continue
                # 3. Check for Thai combining characters out of place (floating tone marks or sara)
                if re.match(r'^[\u0e30-\u0e3a\u0e47-\u0e4e]', s):
                    corrupted_count += 1
                    continue

    if total_cells == 0:
        return False

    return corrupted_count >= 2 or (corrupted_count / max(1, total_cells)) > 0.02

def extract_tables_with_ocr(input_path, page_indices=None, ocr_lang="tha+eng"):
    """
    Extracts tabular data using PyMuPDF high-DPI page rendering and Tesseract OCR with TSV coordinates.
    Reconstructs tables even when PDF fonts lack /ToUnicode or contain backwards stream characters.
    """
    try:
        import fitz
    except ImportError:
        sys.stderr.write("fitz (PyMuPDF) required for OCR table extraction.\n")
        return []

    try:
        doc = fitz.open(input_path)
    except Exception as e:
        sys.stderr.write(f"Cannot open PDF with fitz: {e}\n")
        return []

    total = len(doc)
    target_indices = page_indices if page_indices is not None else list(range(total))
    target_indices = [p for p in target_indices if 0 <= p < total]

    pages_data = []
    dpi = 200
    scale = dpi / 72.0  # 200 / 72 = 2.777778

    with tempfile.TemporaryDirectory() as tmp_dir:
        for p_idx in target_indices:
            page = doc[p_idx]
            pix = page.get_pixmap(dpi=dpi)
            img_path = os.path.join(tmp_dir, f"page_{p_idx}.png")
            pix.save(img_path)

            out_base = os.path.join(tmp_dir, f"ocr_{p_idx}")
            cmd = ['tesseract', img_path, out_base, '-l', ocr_lang, '--psm', '6', 'tsv']
            try:
                subprocess.run(cmd, capture_output=True, timeout=90)
            except Exception as e:
                sys.stderr.write(f"Tesseract OCR psm 6 error: {e}\n")

            tsv_path = f"{out_base}.tsv"
            words = []
            if os.path.exists(tsv_path):
                with open(tsv_path, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.DictReader(f, delimiter='\t', quoting=csv.QUOTE_NONE)
                    for row in reader:
                        if row.get('level') == '5':
                            txt = clean_thai_text(row.get('text', ''))
                            if txt:
                                try:
                                    left = int(row.get('left', 0))
                                    top = int(row.get('top', 0))
                                    width = int(row.get('width', 0))
                                    height = int(row.get('height', 0))
                                    conf = float(row.get('conf', 0))
                                    if conf >= 0:
                                        words.append({
                                            'left': left,
                                            'top': top,
                                            'width': width,
                                            'height': height,
                                            'cx': left + width / 2.0,
                                            'cy': top + height / 2.0,
                                            'conf': conf,
                                            'text': txt
                                        })
                                except (ValueError, TypeError):
                                    pass

            # If psm 6 didn't find words, fallback to psm 3
            if not words:
                cmd3 = ['tesseract', img_path, out_base, '-l', ocr_lang, '--psm', '3', 'tsv']
                try:
                    subprocess.run(cmd3, capture_output=True, timeout=90)
                    if os.path.exists(tsv_path):
                        with open(tsv_path, 'r', encoding='utf-8', errors='ignore') as f:
                            reader = csv.DictReader(f, delimiter='\t', quoting=csv.QUOTE_NONE)
                            for row in reader:
                                if row.get('level') == '5':
                                    txt = clean_thai_text(row.get('text', ''))
                                    if txt:
                                        try:
                                            left = int(row.get('left', 0))
                                            top = int(row.get('top', 0))
                                            width = int(row.get('width', 0))
                                            height = int(row.get('height', 0))
                                            conf = float(row.get('conf', 0))
                                            if conf >= 0:
                                                words.append({
                                                    'left': left,
                                                    'top': top,
                                                    'width': width,
                                                    'height': height,
                                                    'cx': left + width / 2.0,
                                                    'cy': top + height / 2.0,
                                                    'conf': conf,
                                                    'text': txt
                                                })
                                        except (ValueError, TypeError):
                                            pass
                except Exception:
                    pass

            if not words:
                pages_data.append([])
                continue

            page_rows = []
            table_handled = False

            # Method A: Check if page has vector table cells detected by PyMuPDF
            try:
                tabs = page.find_tables()
                if tabs and tabs.tables:
                    for tab in tabs:
                        if hasattr(tab, 'cells') and tab.cells and hasattr(tab, 'row_count') and hasattr(tab, 'col_count'):
                            num_rows = tab.row_count
                            num_cols = tab.col_count
                            grid = [["" for _ in range(num_cols)] for _ in range(num_rows)]
                            for cell_idx, cell_bbox in enumerate(tab.cells):
                                if cell_bbox:
                                    r = cell_idx // num_cols
                                    c = cell_idx % num_cols
                                    if r < num_rows and c < num_cols:
                                        x0, y0, x1, y1 = cell_bbox
                                        img_x0, img_y0, img_x1, img_y1 = x0 * scale, y0 * scale, x1 * scale, y1 * scale
                                        in_cell = [w for w in words if img_x0 - 6 <= w['cx'] <= img_x1 + 6 and img_y0 - 6 <= w['cy'] <= img_y1 + 6]
                                        if in_cell:
                                            in_cell.sort(key=lambda w: (w['top'], w['left']))
                                            grid[r][c] = " ".join(w['text'] for w in in_cell)
                            if any(any(c.strip() for c in row) for row in grid):
                                page_rows.extend(grid)
                                page_rows.append([])
                                table_handled = True
            except Exception as e:
                sys.stderr.write(f"PyMuPDF vector table cell mapping notice: {e}\n")

            if table_handled and page_rows:
                pages_data.append(page_rows)
                continue

            # Method B: Coordinate & Gap Clustering
            words.sort(key=lambda w: (w['top'], w['left']))
            lines = []
            for w in words:
                w_cy = w['cy']
                placed = False
                for line in lines:
                    if abs(w_cy - line['cy']) <= max(12.0, line['height'] * 0.45):
                        line['words'].append(w)
                        line['top'] = min(line['top'], w['top'])
                        line['bottom'] = max(line['bottom'], w['top'] + w['height'])
                        line['height'] = line['bottom'] - line['top']
                        line['cy'] = (line['top'] + line['bottom']) / 2.0
                        placed = True
                        break
                if not placed:
                    lines.append({
                        'top': w['top'],
                        'bottom': w['top'] + w['height'],
                        'height': w['height'],
                        'cy': w['cy'],
                        'words': [w]
                    })

            lines.sort(key=lambda l: l['top'])

            line_segments = []
            for line in lines:
                line['words'].sort(key=lambda w: w['left'])
                segments = []
                curr = []
                for w in line['words']:
                    if not curr:
                        curr.append(w)
                    else:
                        prev = curr[-1]
                        gap = w['left'] - (prev['left'] + prev['width'])
                        if gap > 24:
                            seg_text = " ".join(cw['text'] for cw in curr).strip()
                            segments.append({
                                'text': seg_text,
                                'x0': curr[0]['left'],
                                'x1': prev['left'] + prev['width'],
                                'cx': (curr[0]['left'] + prev['left'] + prev['width']) / 2.0
                            })
                            curr = [w]
                        else:
                            curr.append(w)
                if curr:
                    seg_text = " ".join(cw['text'] for cw in curr).strip()
                    segments.append({
                        'text': seg_text,
                        'x0': curr[0]['left'],
                        'x1': curr[-1]['left'] + curr[-1]['width'],
                        'cx': (curr[0]['left'] + curr[-1]['left'] + curr[-1]['width']) / 2.0
                    })
                line_segments.append(segments)

            all_x0 = sorted([seg['x0'] for segs in line_segments for seg in segs])
            col_anchors = []
            if all_x0:
                for x in all_x0:
                    matched = False
                    for col in col_anchors:
                        if abs(x - col['mean']) < 35:
                            col['vals'].append(x)
                            col['mean'] = sum(col['vals']) / len(col['vals'])
                            matched = True
                            break
                    if not matched:
                        col_anchors.append({'mean': x, 'vals': [x]})

                col_anchors.sort(key=lambda c: c['mean'])
                min_freq = max(1, int(len(line_segments) * 0.08)) if len(line_segments) >= 10 else 1
                valid_cols = [c['mean'] for c in col_anchors if len(c['vals']) >= min_freq]
                valid_cols.sort()
            else:
                valid_cols = []

            for segs in line_segments:
                if not segs:
                    continue
                if not valid_cols or len(valid_cols) <= 1:
                    page_rows.append([s['text'] for s in segs])
                else:
                    row_data = ["" for _ in range(len(valid_cols))]
                    for s in segs:
                        best_col = min(range(len(valid_cols)), key=lambda i: abs(s['x0'] - valid_cols[i]))
                        if row_data[best_col]:
                            row_data[best_col] += " " + s['text']
                        else:
                            row_data[best_col] = s['text']
                    while row_data and not row_data[-1].strip():
                        row_data.pop()
                    if row_data:
                        page_rows.append(row_data)

            pages_data.append(page_rows)

    doc.close()
    return pages_data

def extract_from_pdf(input_path, table_mode="auto", page_indices=None, force_ocr=False):
    """
    Extract tabular structures from PDF using pdfplumber, PyMuPDF, or pdftotext.
    table_mode: 'auto' | 'lattice' | 'stream'
    page_indices: list of 0-based page numbers to extract
    force_ocr: if True, directly use Thai OCR Table extraction
    """
    if force_ocr:
        sys.stderr.write("Running Thai OCR Table Engine...\n")
        return extract_tables_with_ocr(input_path, page_indices=page_indices)

    pages_data = []

    # Method 1: pdfplumber (highest quality table detection)
    try:
        import pdfplumber
        with pdfplumber.open(input_path) as pdf:
            target_indices = page_indices if page_indices is not None else range(len(pdf.pages))
            for p_idx in target_indices:
                if p_idx >= len(pdf.pages):
                    continue
                page = pdf.pages[p_idx]
                page_rows = []

                # Configure table extraction strategy
                table_settings = {}
                if table_mode == "lattice":
                    table_settings = {
                        "vertical_strategy": "lines",
                        "horizontal_strategy": "lines",
                        "snap_tolerance": 3,
                        "join_tolerance": 3,
                    }
                elif table_mode == "stream":
                    table_settings = {
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "snap_tolerance": 5,
                    }

                tables = page.extract_tables(table_settings=table_settings) if table_settings else page.extract_tables()

                # If lattice yielded nothing in auto mode, fallback to text/stream
                if not tables and table_mode == "auto":
                    tables = page.extract_tables(table_settings={"vertical_strategy": "text", "horizontal_strategy": "text"})

                if tables:
                    for table in tables:
                        for row in table:
                            page_rows.append([cell if cell is not None else "" for cell in row])
                        page_rows.append([])
                else:
                    text = page.extract_text()
                    if text:
                        for line in text.splitlines():
                            cols = re.split(r'\s{2,}|\t', line.strip())
                            page_rows.append(cols)
                pages_data.append(page_rows)

        if any(len(p) > 0 for p in pages_data):
            if is_text_corrupted(pages_data):
                sys.stderr.write("Detected corrupted Thai font / (cid:) encoding. Switching to Thai OCR Table Engine...\n")
                ocr_data = extract_tables_with_ocr(input_path, page_indices=page_indices)
                if ocr_data and any(len(p) > 0 for p in ocr_data):
                    return ocr_data
            return pages_data
    except ImportError:
        pass
    except Exception as e:
        sys.stderr.write(f"pdfplumber notice: {e}\n")

    # Method 2: PyMuPDF (fitz) - table extraction
    try:
        import fitz
        doc = fitz.open(input_path)
        pages_data = []
        target_indices = page_indices if page_indices is not None else range(len(doc))

        for p_idx in target_indices:
            if p_idx >= len(doc):
                continue
            page = doc[p_idx]
            page_rows = []
            has_tables = False
            try:
                fitz_strategy = "lines" if table_mode == "lattice" else "lines_strict" if table_mode == "lattice" else None
                tabs = page.find_tables(strategy=fitz_strategy) if fitz_strategy else page.find_tables()
                if tabs and tabs.tables:
                    has_tables = True
                    for tab in tabs:
                        for row in tab.extract():
                            page_rows.append([cell if cell is not None else "" for cell in row])
                        page_rows.append([])
            except Exception:
                has_tables = False

            if not has_tables or not page_rows:
                text = page.get_text("text")
                if text:
                    for line in text.splitlines():
                        cols = re.split(r'\s{2,}|\t', line.strip())
                        page_rows.append(cols)
            pages_data.append(page_rows)

        if any(len(p) > 0 for p in pages_data):
            if is_text_corrupted(pages_data):
                sys.stderr.write("Detected corrupted Thai font / (cid:) encoding in fitz output. Switching to Thai OCR Table Engine...\n")
                ocr_data = extract_tables_with_ocr(input_path, page_indices=page_indices)
                if ocr_data and any(len(p) > 0 for p in ocr_data):
                    return ocr_data
            return pages_data
    except ImportError:
        pass
    except Exception as e:
        sys.stderr.write(f"fitz notice: {e}\n")

    # Method 3: pdftotext -layout (system command from poppler-utils)
    try:
        res = subprocess.run(['pdftotext', '-layout', input_path, '-'], capture_output=True, text=True, timeout=60)
        if res.returncode == 0 and res.stdout:
            raw_pages = res.stdout.split('\x0c')
            pages_data = []
            target_indices = page_indices if page_indices is not None else range(len(raw_pages))
            for p_idx in target_indices:
                if p_idx >= len(raw_pages):
                    continue
                raw_page = raw_pages[p_idx]
                page_rows = []
                for line in raw_page.splitlines():
                    trimmed = line.strip()
                    if trimmed:
                        cols = re.split(r'\s{2,}|\t', trimmed)
                        page_rows.append(cols)
                if page_rows:
                    pages_data.append(page_rows)
            if pages_data:
                if is_text_corrupted(pages_data):
                    sys.stderr.write("Detected corrupted Thai font / (cid:) encoding in pdftotext output. Switching to Thai OCR Table Engine...\n")
                    ocr_data = extract_tables_with_ocr(input_path, page_indices=page_indices)
                    if ocr_data and any(len(p) > 0 for p in ocr_data):
                        return ocr_data
                return pages_data
    except Exception as e:
        sys.stderr.write(f"pdftotext notice: {e}\n")

    # Final fallback if vector extraction failed completely
    sys.stderr.write("Vector table extraction yielded no content. Running Thai OCR Table Engine...\n")
    return extract_tables_with_ocr(input_path, page_indices=page_indices)

def get_pdf_total_pages(input_path):
    try:
        import fitz
        doc = fitz.open(input_path)
        return len(doc)
    except Exception:
        pass
    try:
        import pdfplumber
        with pdfplumber.open(input_path) as pdf:
            return len(pdf.pages)
    except Exception:
        pass
    return 1

def main():
    parser = argparse.ArgumentParser(description="Advanced PDF to Excel converter")
    parser.add_argument("input_pdf", help="Path to input PDF file")
    parser.add_argument("output_xlsx", help="Path to output XLSX file")
    parser.add_argument("--table-mode", dest="table_mode", default="auto", choices=["auto", "lattice", "stream"],
                        help="Table detection strategy: auto, lattice (bordered), stream (whitespace)")
    parser.add_argument("--sheet-mode", dest="sheet_mode", default="single", choices=["single", "multiple"],
                        help="Sheet structure: single (one continuous sheet) or multiple (sheet per page)")
    parser.add_argument("--pages", dest="pages", default="all",
                        help="Pages to convert: 'all' or comma/range (e.g. '1,3-5')")
    parser.add_argument("--ocr", action="store_true", help="Force Thai OCR table extraction engine")

    args = parser.parse_args()

    if not os.path.isfile(args.input_pdf):
        sys.stderr.write(f"Input file not found: {args.input_pdf}\n")
        sys.exit(1)

    total_pages = get_pdf_total_pages(args.input_pdf)
    target_page_indices = parse_page_ranges(args.pages, total_pages)

    pages_data = extract_from_pdf(args.input_pdf, table_mode=args.table_mode, page_indices=target_page_indices, force_ocr=args.ocr)
    if not pages_data:
        pages_data = [[["ไม่พบข้อมูลตารางในเอกสาร PDF หรือเป็นไฟล์รูปภาพสแกน"]]]

    save_to_excel(pages_data, args.output_xlsx, sheet_mode=args.sheet_mode)

    if os.path.isfile(args.output_xlsx) and os.path.getsize(args.output_xlsx) > 0:
        print(f"Success: {args.output_xlsx}")
        sys.exit(0)
    else:
        sys.stderr.write("Failed to create xlsx file\n")
        sys.exit(1)

if __name__ == '__main__':
    main()
